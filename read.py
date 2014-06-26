#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl import load_workbook
import re
import sys

# workbooks to process
files = [
    "Chinavitae_1_500.xlsm",
    "Chinavitae_501_1000.xlsm",
    "Chinavitae_1001_1500.xlsm",
    "Chinavitae_1501_2000.xlsm",
    "Chinavitae_2001_2500.xlsm",
    "Chinavitae_2501_3000.xlsm",
    "Chinavitae_3001_3500.xlsm",
    "Chinavitae_3501_4000.xlsm",
    "Chinavitae_4001_4509.xlsm"
]

# annoying seperators
dot = u'\u00b7'
dash = u'\u2014'
emph = u'\u2022'
dot2 = u'\u2027'

seps = (u'.', dot, dash, emph, dot2)

# regex pattern matching all ascii characters
asciiPattern = ur'[%s]+' % ''.join(chr(i) for i in range(32, 127))
# regex pattern matching all common Chinese characters and seporators
chinesePattern = ur'[\u4e00-\u9fff. %s]+' % (''.join(seps))


def split_name(name):
    """Split [English name, Chinese name].

        If one of them is missing, None will be returned instead.
    Usage:
        engName, chName = split_name(name)
    """
    matches = re.match('(%s) (%s)' % (asciiPattern, chinesePattern), name)
    if matches:  # English name + Chinese name
        return matches.group(1).strip(), matches.group(2).strip()
    else:
        matches = re.findall('(%s)' % (chinesePattern), name)
        matches = ''.join(matches).strip()
        if matches:  # Chinese name only
            return None, matches
        else:  # English name only
            matches = re.findall('(%s)' % (asciiPattern), name)
            return ''.join(matches).strip(), None


def get_clean_ch_name(chName):
    """Remove annoying seperators from the Chinese name of minorities.

    Usage:
        cleanName = get_clean_ch_name(chName)
    """
    cleanName = chName
    for sep in seps:
        cleanName = cleanName.replace(sep, u' ')
    return cleanName


def split_ch_name(chName):
    """Split the Chinese name into first name and last name.

        * If the name is XY or XYZ, X will be returned as the first name.
        * If the name is WXYZ, WX will be returned as the first name.
        * If the name is ...WXYZ, the whole name will be returned
          as the last name.
        * If the name is ..ABC * XYZ..., the part before the seperator
          will be returned as the first name.
    Usage:
        chFirstName, chLastName = split_ch_name(chName)
    """
    if len(chName) < 4:  # XY or XYZ
        chFirstName = chName[0]
        chLastName = chName[1:]
    elif len(chName) == 4:  # WXYZ
        chFirstName = chName[:2]
        chLastName = chName[2:]
    else:  # longer
        cleanName = get_clean_ch_name(chName)
        nameParts = cleanName.split()
        print u' '.join(nameParts)
        if len(nameParts) < 2:  # ...WXYZ
            return None, nameParts[0]
        chFirstName, chLastName = nameParts[:2]  # ..ABC * XYZ...
    return chFirstName, chLastName


def split_eng_name(engName):
    """Split the English name into first name and last name.

        * If there is no spaces, the whole name is returned as the first name
        * If the are spaces, they will be split using the space
    Usage:
        engFirstName, engLastName = split_eng_name(engName)
    """
    nameParts = engName.split()
    if len(nameParts) < 2:
        return nameParts[0], None
    engFirstName, engLastName = nameParts[:2]
    return engFirstName, engLastName


def guess_bio(bioLine):
    """Guessing the person's gender and nationality from the biography.
        
    Usage:
        gender, nationality = split_ch_name(bioLine)
    """
    # guessing gender
    gender = None
    if re.search(ur'\bmale\b', bioLine):
        gender = 'Male'
    elif re.search(ur'\bfemale\b', bioLine):
        gender = 'Female'
    # guessing nationality
    nationality = re.findall(ur'(\w+) nationality', bioLine)
    nationality = nationality[0] if len(nationality) > 0 else None
    return gender, nationality


def guess_career(line):
    """Guessing the person's role and entity from the career data.
            
    Usage:
        role, entity = guess_career(line)
    """
    role = entity = None
    # guess role and entity
    role, temp, entity = line[1].partition(', ')
    return role, entity


def guess_travel(line):
    """Guessing the nature of a person's trip from the travel data.
    
    Usage:
        nature = guess_travel(line)
    """
    nature = 'News'
    # guess nature
    if line[2].startswith(u'Travelled to') or line[2].startswith(u'Was in'):
        nature = 'Location'
    elif line[2].startswith(emph):
        nature = 'Meeting'
    return nature


def main():
    """ This script should be run like:
            python read.py 0
        where 0 can be replaced by other numbers indicating the workbook to use.
    """
    # the the command line argument
    if sys.argv[1] and sys.argv[1].isdigit():
        filename = files[int(sys.argv[1])]
    else:
        return

    # input workbook
    print "loading", filename, "......."
    inwb = load_workbook(filename)

    # output workbook
    outwb = Workbook()

    # create sheets and heads
    travelSheet = outwb.create_sheet(0, 'travel')
    travelSheet.append(
        ('IndividualID', 'Date', 'Type', 'Description', 'Nature'))

    careerSheet = outwb.create_sheet(0, 'career')
    careerSheet.append(('IndividualID', 'Derived', 'Date', 'SourceInformation',
                        'RoleTitle', 'EntityName', 'Comments'))

    infoSheet = outwb.create_sheet(0, 'info')
    infoSheet.append(('BioLastRevised', 'CareerLastUpdate',
                      'ChineseName', 'EnglishFirstName', 'EnglishLastName',
                      'ChineseFirstName', 'ChineseLastName', 
                      'YearOfBirth', 'BirthPlace', 'Gender',
                      'Nationality', 'Biography'))

    # Loop through the CVs in this woorkbook
    for sheetName in inwb.get_sheet_names():
        if not sheetName.isdigit():
            continue
        sheet = inwb[sheetName]
        print "------------------------------------------------------"
        print "Processing sheet", sheetName, "........"
        print ' '

        # Only one column in this sheet, it is invalid, skip
        if len(sheet.columns) < 2:
            continue

        colA, colB = sheet.columns[:2]

        # No comments in this sheet, create an empty column
        if len(sheet.columns) <= 2:
            alen = len(colA)
            for i in range(1, alen):
                sheet.cell('C%s' % (i)).value = None

        colC = sheet.columns[2]

        # intialize for each CV
        revisedTime = updatedTime = birthYear = engName = chName = None
        birthPlace = bioLineIdx = careerIdx = travelIdx = compIdx = None
        engFirstName = engLastName = chFirstName = chLastName = None
        bioLine = gender = nationality = None

        # process column A to get all available fields and indexes
        for idx, cell in enumerate(colA):
            if unicode(cell.value).startswith(u'Biography Revised:'):
                revisedTime = cell.value.partition(':')[-1].strip()
            if unicode(cell.value).startswith(u'Career Data Updated:'):
                updatedTime = cell.value.partition(':')[-1].strip()
            if unicode(cell.value).startswith(u'Born:'):
                birthYear = cell.value.partition(':')[-1].strip()
            if unicode(cell.value).startswith(u'PHOTO:'):
                engName, chName = split_name(colA[idx + 1].value)
            if unicode(cell.value).startswith(u'Birthplace:'):
                birthPlace = cell.value.partition(':')[-1].strip()
            if unicode(cell.value) == u'Biography':
                bioLineIdx = idx
            if unicode(cell.value) == u'Career Data':
                careerIdx = idx
            if unicode(cell.value).startswith(u'Recent Travel'):
                travelIdx = idx
            if unicode(cell.value) == u'Compare':
                compIdx = idx

        # English name
        if engName:
            engFirstName, engLastName = split_eng_name(engName)

        # Chinese name
        if chName:
            chFirstName, chLastName = split_ch_name(chName)

        # Biography
        if bioLineIdx:
            # Merge all lines
            bioLine = colA[bioLineIdx + 2].value
            if not bioLine:
                bioLine = ''
            else:
                bioLine = unicode(bioLine)
            i = 3
            nextLine = colA[bioLineIdx + i].value
            while nextLine != 'Career Data' and nextLine != 'Recent Travel':
                if nextLine:
                    bioLine += ' ' + unicode(nextLine)
                i = i + 1
                nextLine = colA[bioLineIdx + i].value
            # guess gender and nationality
            gender, nationality = guess_bio(bioLine)

        # Info to go with the profile
        cv = {
            'revisedTime': revisedTime,
            'updatedTime': updatedTime,
            'chName': chName,
            'engFirstName': engFirstName,
            'engLastName': engLastName,
            'chFirstName': chFirstName,
            'chLastName': chLastName,
            'birthYear': birthYear,
            'birthPlace': birthPlace,
            'gender': gender,
            'nationality': nationality,
            'bioLine': bioLine,
            'id': sheetName
        }

        infoSheet.append((cv['revisedTime'], cv['updatedTime'],
                          cv['chName'], cv['engFirstName'], cv[
                          'engLastName'], cv['chFirstName'],
                          cv['chLastName'], cv['birthYear'], cv[
                          'birthPlace'], cv['gender'],
                          cv['nationality'], cv['bioLine']))

        # used to output to console
        name = get_clean_ch_name(chName) if chName else engName
        print "Insert profile for", cv['id'], name

        # get the career info
        if careerIdx:
            if travelIdx:  # followed by the  travel block
                print "careerIdx, travelIdx", careerIdx, travelIdx
                cv['careerData'] = [(colA[i].value, colB[i].value, colC[i].value)
                                    for i in range(careerIdx + 2, travelIdx - 1)]
            else:  # it's the last useful block
                print "careerIdx, compIdx", careerIdx, compIdx
                cv['careerData'] = [(colA[i].value, colB[i].value, colC[i].value)
                                    for i in range(careerIdx + 2, compIdx)]

            # Compile the lines
            for idx, line in enumerate(cv['careerData']):
                # empty line, skip
                if not line[0] and not line[1] and not line[2]:
                    continue

                # guess role and entity
                role, entity = guess_career(line)

                # output to the result
                careerSheet.append((cv['id'], None, line[0], line[1],
                                    role, entity, line[2]))
            print "Inserted career data for", cv['id'], name

        # get the travel info
        if travelIdx:
            # it's usually the last block
            cv['travelData'] = [[colA[i].value, colB[i].value, colC[i].value]
                                for i in range(travelIdx + 2, compIdx)]

            for idx, line in enumerate(cv['travelData']):
                # empty line, skip
                if not line[0] and not line[1] and not line[2]:
                    continue

                # ignore some annoying messages e.g. Email scan, how many
                # more...
                if type(line[0]) is unicode:
                    break

                # guess the nature
                nature = guess_travel(line)

                # pull down the empty fields
                if not line[0]:
                    line[0] = cv['travelData'][idx - 1][0] if idx > 0 else None
                if not line[1]:
                    line[1] = cv['travelData'][idx - 1][1] if idx > 0 else None

                # output to result
                travelSheet.append(
                    (cv['id'], line[0], line[1], line[2], nature))
            print "Inserted travel data for", cv['id'], name

    # save the result
    print "saving into out%s.xlsx...." % sys.argv[1]
    outwb.save("out%s.xlsx" % sys.argv[1])


if __name__ == "__main__":
    main()
